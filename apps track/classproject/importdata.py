

import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
import django
django.setup()
import click
import openpyxl
from django.db import models

from onlineapp.models import *
@click.group()
def dboperations():
    click.echo("hiii")
    pass

@dboperations.command('importdata',short_help='imports the data from the corresponding excel sheets')
@click.argument('files')
def importdata(files):
    wb=openpyxl.load_workbook(files)
    sheet=wb['Colleges']
    row=sheet.max_row
    column=sheet.max_column
    for r in range(2,row+1):
        row_data=[]
        for c in range(1,column+1):
            e=sheet.cell(row=r,column=c)
            row_data.append(e.value)
        c=College(name=row_data[0],location=row_data[2],acronym=row_data[1],contact=row_data[3])
        #mycursor.execute(query,tuple(row_data))
        c.save()



if __name__=='__main__':
    dboperations()